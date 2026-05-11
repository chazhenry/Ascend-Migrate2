import csv
import zipfile
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.job import Job
from app.services.jobs import append_job_log, save_artifact
from app.services.pipeline import acquisition_workdir, finalize_stage, get_acquisition, mark_stage_running


async def run_stage(acquisition_id: UUID, job_id: UUID, db: AsyncSession) -> None:
    acquisition = await get_acquisition(db, acquisition_id)
    await mark_stage_running(db, acquisition)
    job = (await db.execute(select(Job).where(Job.id == job_id))).scalar_one()

    workdir = acquisition_workdir(acquisition_id)
    outputs_dir = workdir / "etl_output"
    excel_path = workdir / "cch_axcess_output.xlsx"
    zip_path = workdir / "cch_flat_files.zip"

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    csv_paths = sorted(outputs_dir.glob("*.csv")) if outputs_dir.exists() else []
    if not csv_paths:
        sheet = workbook.create_sheet(title="Summary")
        sheet.append(["status", "No ETL outputs were available when output generation ran."])

    for csv_path in csv_paths:
        title = csv_path.stem[:31]
        sheet = workbook.create_sheet(title=title)
        with csv_path.open("r", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            for row in reader:
                sheet.append(row)
        await append_job_log(db, job, f"Added {csv_path.name} to Excel workbook.")

    workbook.save(excel_path)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zip_handle:
        for csv_path in csv_paths:
            zip_handle.write(csv_path, arcname=csv_path.name)
        if not csv_paths:
            placeholder = workdir / "README.txt"
            placeholder.write_text("No ETL outputs were available when the flat-file bundle was generated.\n", encoding="utf-8")
            zip_handle.write(placeholder, arcname="README.txt")

    await save_artifact(db, acquisition_id, 7, "cch_excel_output", file_path=str(excel_path))
    await save_artifact(db, acquisition_id, 7, "cch_flat_files", file_path=str(zip_path))
    await append_job_log(db, job, "Generated Excel and flat-file outputs.")
    await finalize_stage(db, acquisition, 7, "complete")
